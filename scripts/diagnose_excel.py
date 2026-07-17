"""Excel yükleme teşhisi — kök neden analizi için."""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "backend"))

from openpyxl import load_workbook

from app.repositories.excel_repository import (
    ExcelStockRepository,
    _normalize_col,
    _resolve_columns,
)

SEARCH_TERMS = ["5329057", "MOVA", "2606160168", "A01E"]


def inspect_raw(path: Path) -> None:
    print("=" * 70)
    print(f"DOSYA: {path.name} ({path.stat().st_size} byte)")
    print("=" * 70)

    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"Sayfalar: {wb.sheetnames}")
    ws = wb.active
    print(f"Aktif sayfa: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))
    total = len(rows)
    print(f"\n1) TOPLAM SATIR (sayfa): {total}")

    print("\n--- Ilk 10 satir (ham) ---")
    for i, row in enumerate(rows[:10], start=1):
        print(f"  Satir {i}: {row}")

    print("\n--- Son 10 satir (ham) ---")
    for i, row in enumerate(rows[-10:], start=total - 9):
        print(f"  Satir {i}: {row}")

    # Header tespiti — ilk 5 satirda "Etiket" ve "Depo" ara
    print("\n2) BASLIK SATIRI TESPITI")
    header_row_idx = None
    for i, row in enumerate(rows[:20], start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        norm = {_normalize_col(c) for c in cells if c}
        if "etiket" in norm and ("depo" in norm or "raf" in norm):
            header_row_idx = i
            print(f"  Baslik adayi satir {i}: {cells}")
            break
    if header_row_idx is None:
        print("  UYARI: Ilk 20 satirda Etiket+Depo basligi bulunamadi!")
        header_row_idx = 1
        cells = [str(c).strip() if c is not None else "" for c in rows[0]]
        print(f"  Varsayilan satir 1: {cells}")

    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx - 1]]
    print(f"\n3) KULLANILAN BASLIK (satir {header_row_idx}): {header}")

    try:
        col_map = _resolve_columns(header)
        print(f"\n4) SUTUN ESLESMESI (mevcut kod — satir 1 varsayimi):")
        for k, v in col_map.items():
            idx = header.index(v)
            print(f"  {k:10} -> '{v}' (index {idx}, harf {chr(65+idx) if idx < 26 else '?'})")
    except Exception as e:
        print(f"\n4) SUTUN ESLESMESI HATASI (satir 1 ile): {e}")
        col_map = {}

    # 5329057 / MOVA arama
    print("\n5) '5329057' / 'MOVA' ARAMASI (tum sayfa)")
    for i, row in enumerate(rows, start=1):
        for j, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell)
            if "5329057" in s or "MOVA" in s.upper():
                col = header[j] if j < len(header) else f"col{j}"
                print(f"  Satir {i}, kolon {j} ('{col}'): {cell!r}")

    wb.close()


def inspect_repository(path: Path) -> None:
    print("\n" + "=" * 70)
    print("REPOSITORY YUKLEME (load_from_excel — satir 1 = baslik)")
    print("=" * 70)

    repo = ExcelStockRepository()
    try:
        meta = repo.load_from_excel(str(path))
    except Exception as e:
        print(f"YUKLEME HATASI: {e}")
        return

    print(f"Toplam satir: {meta['row_count']}")
    print(f"Etiket sayisi: {meta['etiket_count']}")
    print(f"Raf sayisi: {meta['shelf_count']}")
    print(f"Miktar sutunu: {meta.get('miktar_column')}")
    print(f"Stok sutunu: {meta.get('stok_column')}")
    print(f"Sutunlar: {meta.get('columns')}")

    shelves = repo.get_shelves()
    print(f"\n6) RAFLAR ({len(shelves)}): {shelves[:20]}{'...' if len(shelves) > 20 else ''}")
    print("   Veri yapisi: ExcelStockRepository._shelf_lines (dict[shelf][line_id] -> item)")

    a01e = repo.get_shelf_items("A01E")
    print(f"\n7) A01E RAFI — {len(a01e)} satir:")
    for i, item in enumerate(a01e, start=1):
        extra = item.get("extra", {})
        stok = extra.get("Stok No", "")
        tanim = extra.get("Tanım", extra.get("Tanim", ""))
        print(
            f"  [{i}] etiket={item['etiket']} stok={stok} tanim={tanim!r} "
            f"depo=A01E miktar={item['expected']} line_id={item['line_id']}"
        )

    if not a01e:
        print("  A01E BOS! Benzer raf adlari:")
        for s in shelves:
            if "A01" in s.upper():
                print(f"    {s}: {len(repo.get_shelf_items(s))} satir")


def main() -> None:
    uploads = ROOT / "backend" / "uploads"
    files = sorted(uploads.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        print("uploads/ klasorunde xlsx yok")
        return

    for path in files[:2]:
        inspect_raw(path)
        inspect_repository(path)
        print("\n")


if __name__ == "__main__":
    main()
