from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.config import settings
from app.models.domain import CountTrackingStatus, ItemStatus
from app.services.count_service import CountService


class ReportService:
    def __init__(self, count_service: CountService) -> None:
        self.count = count_service

    @staticmethod
    def _misplacement_message(c: dict[str, Any]) -> str:
        status = c.get("status", "")
        scanned = c.get("scanned_shelf") or ""
        correct = c.get("correct_shelf") or ""
        if status == "Boş raf bilgisi":
            return f"Excel'de raf/depo boş — {scanned} rafında okutuldu"
        if status == "Raf bulunamadı":
            return f"Excel'de kayıt yok — {scanned} rafında okutuldu"
        if correct:
            return f"{correct} rafına ait — burada ({scanned}) yanlış okutuldu"
        return f"{scanned} rafında okutuldu"

    @staticmethod
    def _misplacement_category(status: str) -> str:
        if status == "Boş raf bilgisi":
            return "Depo boş"
        if status == "Raf bulunamadı":
            return "Excel'de yok"
        return "Raf uyumsuzluğu"

    @staticmethod
    def _not_found_message(row: dict[str, Any]) -> tuple[str, str]:
        status = row.get("tracking_status", "")
        expected = row.get("expected_shelf") or ""
        found = row.get("found_shelf") or ""
        if status == CountTrackingStatus.BULUNAMADI.value:
            return (
                "Gerçek eksik",
                f"Sayım sonunda bulunamadı — {expected} rafında aranan ürün eksik",
            )
        if status == CountTrackingStatus.SONRADAN_BULUNDU.value:
            return (
                "Yanlış lokasyonda bulundu",
                f"Bulunamadı işaretliydi — {expected} yerine {found} rafında bulundu",
            )
        if status == CountTrackingStatus.TEKRAR_BULUNDU.value:
            return (
                "Doğru rafta bulundu",
                f"Bulunamadı işaretliydi — {expected} rafında tekrar bulundu",
            )
        return ("Bulunamadı kaydı", f"Durum: {status}")

    @classmethod
    def build_correction_log(
        cls,
        corrections: list[dict[str, Any]],
        not_found_rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        entries: list[dict[str, Any]] = []

        for c in corrections:
            status = c.get("status", "")
            entries.append(
                {
                    "etiket": c["etiket"],
                    "category": cls._misplacement_category(status),
                    "message": cls._misplacement_message(c),
                    "expected_shelf": c.get("correct_shelf") or "",
                    "found_shelf": c.get("scanned_shelf") or "",
                    "stok_no": "",
                    "product_name": "",
                    "username": c.get("username") or "",
                    "created_at": c.get("created_at") or "",
                }
            )

        for n in not_found_rows:
            category, message = cls._not_found_message(n)
            when = n.get("resolved_at") or n.get("marked_at") or ""
            user = n.get("resolved_by_name") or n.get("marked_by_name") or ""
            entries.append(
                {
                    "etiket": n["etiket"],
                    "category": category,
                    "message": message,
                    "expected_shelf": n.get("expected_shelf") or "",
                    "found_shelf": n.get("found_shelf") or "",
                    "stok_no": n.get("stok_no") or "",
                    "product_name": n.get("product_name") or "",
                    "username": user,
                    "created_at": when,
                }
            )

        def sort_key(item: dict[str, Any]) -> str:
            return str(item.get("created_at") or "")

        entries.sort(key=sort_key, reverse=True)
        for idx, entry in enumerate(entries, start=1):
            entry["row_no"] = idx
        return entries

    async def build_summary(self, session: dict) -> dict[str, Any]:
        session_id = session["id"]
        all_items: list[dict] = []
        complete, short, over, pending = [], [], [], []

        for shelf in self.count.stock.get_shelves():
            detail = self.count.get_shelf_detail(shelf)
            for item in detail["items"]:
                row = {
                    "etiket": item.etiket,
                    "shelf": shelf,
                    "expected": item.expected,
                    "scanned": item.scanned,
                    "status": item.status.value,
                    "extra": item.extra,
                }
                all_items.append(row)
                if item.status == ItemStatus.COMPLETE:
                    complete.append(row)
                elif item.status == ItemStatus.SHORT:
                    short.append(row)
                elif item.status == ItemStatus.OVER:
                    over.append(row)
                else:
                    pending.append(row)

        unknown = await self.count.sessions.get_unknown_items(session_id)
        unassigned = await self.count.sessions.get_unassigned_found(session_id)
        corrections = await self.count.sessions.get_misplacements(session_id)
        not_found_rows = await self.count.sessions.get_not_found_markings(session_id)
        shelf_stats = self.count.get_all_shelf_summaries()
        scan_events = await self.count.sessions.get_scan_events(session_id)

        still_not_found = [
            r for r in not_found_rows
            if r["tracking_status"] == CountTrackingStatus.BULUNAMADI.value
        ]
        found_after_missing = [
            r for r in not_found_rows
            if r["tracking_status"] == CountTrackingStatus.SONRADAN_BULUNDU.value
        ]
        found_on_correct = [
            r for r in not_found_rows
            if r["tracking_status"] == CountTrackingStatus.TEKRAR_BULUNDU.value
        ]

        correction_log = self.build_correction_log(corrections, not_found_rows)

        started = session.get("started_at")
        ended = session.get("ended_at") or datetime.utcnow().isoformat()
        duration = 0.0
        if started:
            try:
                s = datetime.fromisoformat(started.replace("Z", "+00:00"))
                e = datetime.fromisoformat(ended.replace("Z", "+00:00"))
                duration = (e - s).total_seconds() / 60.0
            except ValueError:
                pass

        total_expected = sum(i["expected"] for i in all_items)
        total_scanned = sum(i["scanned"] for i in all_items)
        perf = round((total_scanned / total_expected * 100), 1) if total_expected else 0.0

        return {
            "session": session,
            "duration_minutes": round(duration, 1),
            "complete": complete,
            "short": short,
            "over": over,
            "pending": pending,
            "unknown": unknown,
            "unassigned_found": unassigned,
            "corrections": corrections,
            "correction_log": correction_log,
            "not_found_markings": not_found_rows,
            "not_found_still_missing": still_not_found,
            "found_after_missing": found_after_missing,
            "found_on_correct_shelf": found_on_correct,
            "shelf_stats": shelf_stats,
            "scan_events": scan_events,
            "total_expected": total_expected,
            "total_scanned": total_scanned,
            "performance_pct": perf,
        }

    def _write_summary_sheet(self, wb: Workbook, summary: dict[str, Any]) -> None:
        ws = wb.active
        ws.title = "Özet"
        session = summary["session"]
        ws.append(["Sayım Raporu Özeti"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append(["Oturum", session.get("name", "")])
        ws.append(["Oturum ID", session.get("id", "")])
        ws.append(["Süre (dk)", summary["duration_minutes"]])
        ws.append(["Performans %", summary["performance_pct"]])
        ws.append([])
        ws.append(["Tam", len(summary["complete"])])
        ws.append(["Eksik", len(summary["short"])])
        ws.append(["Fazla", len(summary["over"])])
        ws.append(["Bekleyen", len(summary["pending"])])
        ws.append(["Düzeltme kaydı", len(summary.get("correction_log", []))])
        ws.append(["Gerçek eksik (bulunamadı)", len(summary.get("not_found_still_missing", []))])
        ws.append(["Sonradan bulunan", len(summary.get("found_after_missing", []))])
        ws.append([])
        ws.append(["Detaylı düzeltme listesi için 'Düzeltmeler' sayfasına bakın."])

    def _write_corrections_sheet(self, wb: Workbook, summary: dict[str, Any]) -> None:
        ws = wb.create_sheet("Düzeltmeler")
        headers = [
            "Sıra",
            "Etiket",
            "Kategori",
            "Bildirim",
            "Beklenen Raf",
            "Okutulan / Bulunan Raf",
            "Stok Kodu",
            "Ürün Adı",
            "Kullanıcı",
            "Tarih",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for entry in summary.get("correction_log", []):
            ws.append([
                entry.get("row_no", ""),
                entry["etiket"],
                entry["category"],
                entry["message"],
                entry.get("expected_shelf") or "",
                entry.get("found_shelf") or "",
                entry.get("stok_no") or "",
                entry.get("product_name") or "",
                entry.get("username") or "",
                entry.get("created_at") or "",
            ])

        ws.column_dimensions["D"].width = 60
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col_idx, width in enumerate([6, 16, 18, 60, 14, 18, 14, 24, 14, 20], start=1):
            letter = get_column_letter(col_idx)
            if letter != "D":
                ws.column_dimensions[letter].width = width

        if ws.max_row == 1:
            ws.append(["—", "—", "—", "Bu sayımda düzeltme kaydı yok.", "", "", "", "", "", ""])

    def export_excel(self, summary: dict[str, Any], filename: str) -> Path:
        wb = Workbook()
        self._write_summary_sheet(wb, summary)
        self._write_corrections_sheet(wb, summary)

        sheets = [
            ("Tam Sayılan", summary["complete"]),
            ("Eksik Sayılan", summary["short"]),
            ("Fazla Sayılan", summary["over"]),
            ("Bekleyen", summary["pending"]),
        ]
        for title, rows in sheets:
            ws = wb.create_sheet(title)
            ws.append(["Etiket", "Raf", "Beklenen", "Okutulan", "Durum"])
            for r in rows:
                ws.append([r["etiket"], r["shelf"], r["expected"], r["scanned"], r["status"]])

        ws_unknown = wb.create_sheet("Excel'de Olmayan")
        ws_unknown.append(["Etiket", "Okutulan", "Raf", "Kullanıcı", "Son Okutma"])
        for u in summary["unknown"]:
            ws_unknown.append([
                u["reference"], u["scanned_qty"], u["shelf"], u["username"], u["last_scan_at"]
            ])

        ws_unassigned = wb.create_sheet("Atanmamış Bulunan")
        ws_unassigned.append(["Etiket", "Bulunduğu Raf", "Miktar", "Durum", "Kullanıcı", "Tarih"])
        for u in summary["unassigned_found"]:
            ws_unassigned.append([
                u["reference"], u["found_shelf"], u["scanned_qty"], u["status"], u["username"], u["counted_at"]
            ])

        ws_shelf = wb.create_sheet("Raf Bazlı")
        ws_shelf.append([
            "Raf", "Toplam Etiket", "Tamamlanan", "Eksik", "Fazla", "Bekleyen",
            "Beklenen Adet", "Okutulan", "Tamamlanma %"
        ])
        for s in summary["shelf_stats"]:
            ws_shelf.append([
                s["shelf"], s["total_etikets"], s["completed_etikets"],
                s["short_etikets"], s["over_etikets"], s["pending_etikets"],
                s["total_expected"], s["total_scanned"], s["completion_pct"],
            ])

        ws_users = wb.create_sheet("Kullanıcı Hareketleri")
        ws_users.append(["Etiket", "Raf", "Tip", "Beklenen", "Okutulan", "Kullanıcı", "Saat"])
        for e in summary["scan_events"]:
            ws_users.append([
                e["reference"], e["shelf"], e["scan_type"],
                e["expected"], e["scanned"], e["username"], e["scanned_at"],
            ])

        out = settings.report_dir / filename
        wb.save(out)
        return out

    def export_pdf(self, summary: dict[str, Any], filename: str) -> Path:
        out = settings.report_dir / filename
        doc = SimpleDocTemplate(str(out), pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story = []

        session = summary["session"]
        story.append(Paragraph(f"<b>Sayım Raporu — {session.get('name', '')}</b>", styles["Title"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph(
            f"Süre: {summary['duration_minutes']} dk | "
            f"Performans: %{summary['performance_pct']} | "
            f"Tam: {len(summary['complete'])} | Eksik: {len(summary['short'])} | "
            f"Fazla: {len(summary['over'])} | "
            f"Düzeltme: {len(summary.get('correction_log', []))} | "
            f"Gerçek eksik: {len(summary.get('not_found_still_missing', []))}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 16))

        corr_data = [["#", "Etiket", "Kategori", "Bildirim", "Kullanıcı", "Tarih"]]
        for entry in summary.get("correction_log", []):
            corr_data.append([
                str(entry.get("row_no", "")),
                entry["etiket"],
                entry["category"],
                entry["message"],
                entry.get("username") or "",
                str(entry.get("created_at") or "")[:19],
            ])
        if len(corr_data) == 1:
            corr_data.append(["—", "—", "—", "Düzeltme kaydı yok", "", ""])

        story.append(Paragraph("<b>Düzeltmeler — Kalem Kalem Bildirimler</b>", styles["Heading2"]))
        story.append(Spacer(1, 8))
        corr_table = Table(corr_data, repeatRows=1, colWidths=[24, 72, 80, 280, 60, 90])
        corr_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(corr_table)
        story.append(Spacer(1, 16))

        data = [["Raf", "Etiket", "Tam", "Eksik", "Fazla", "Beklenen", "Okutulan", "%"]]
        for s in summary["shelf_stats"]:
            data.append([
                s["shelf"], str(s["total_etikets"]), str(s["completed_etikets"]),
                str(s["short_etikets"]), str(s["over_etikets"]),
                str(int(s["total_expected"])), str(int(s["total_scanned"])),
                str(s["completion_pct"]),
            ])

        story.append(Paragraph("<b>Raf Bazlı Özet</b>", styles["Heading2"]))
        story.append(Spacer(1, 8))
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(table)
        doc.build(story)
        return out

    @staticmethod
    def list_report_files(limit: int = 50) -> list[dict[str, Any]]:
        files: list[dict[str, Any]] = []
        for path in sorted(
            settings.report_dir.glob("sayim_raporu_*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )[:limit]:
            stat = path.stat()
            files.append(
                {
                    "filename": path.name,
                    "size_bytes": stat.st_size,
                    "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                }
            )
        return files
