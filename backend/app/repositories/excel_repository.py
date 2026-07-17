"""
Excel RAM cache — tek merkezi veri katmanı.

Excel dosyası yalnızca load_from_excel() ile diskten okunur; sonrasında tüm
sorgular bellek içi indeksler üzerinden O(1) / O(k) çalışır.
"""
from __future__ import annotations

import hashlib
import re
import threading
import time
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from app.core.logging import logger
from app.models.domain import EtiketInfo, StockRow
from app.repositories.interfaces import StockRepository
from app.utils.depo import normalize_depo
from app.utils.etiket import (
    compile_etiket_pattern,
    default_etiket_pattern,
    infer_etiket_format,
    matches_etiket_format,
    normalize_etiket,
)

REQUIRED_COLUMNS = {"etiket", "depo", "miktar"}

COLUMN_ALIASES = {
    "etiket": [
        "etiket",
        "etiket_kodu",
        "label",
        "tag",
        "barkod",
        "qr",
    ],
    "depo": ["depo", "raf", "shelf", "lokasyon", "location"],
}

MIKTAR_ALIASES_PRIORITY = [
    "kalan_mik",
    "kalan_miktar",
    "eksilen_mik",
    "eksilen_miktar",
    "mik",
    "miktar",
    "quantity",
    "qty",
    "adet",
    "stok",
    "miktar_öb",
    "miktar_ob",
    "çek._miktar",
    "cek_miktar",
    "çekilen_miktar",
    "paket_miktarı",
    "paket_miktari",
]

STOK_ALIASES = [
    "stok_no",
    "stok_no.",
    "stokno",
    "stok_kodu",
    "stok_kod",
    "malzeme",
    "material",
]


def _normalize_col(name: str) -> str:
    return str(name).strip().lower().replace(" ", "_")


def _resolve_columns(columns: list[str]) -> dict[str, str]:
    normalized = {_normalize_col(c): c for c in columns}
    mapping: dict[str, str] = {}

    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mapping[canonical] = normalized[alias]
                break

    if "miktar" not in mapping:
        for alias in MIKTAR_ALIASES_PRIORITY:
            if alias in normalized:
                mapping["miktar"] = normalized[alias]
                break

    missing = REQUIRED_COLUMNS - set(mapping.keys())
    if missing:
        labels = {"etiket": "Etiket", "depo": "Depo", "miktar": "Miktar/Kalan Mik"}
        missing_tr = ", ".join(labels.get(m, m) for m in sorted(missing))
        raise ValueError(
            f"Excel dosyasında zorunlu sütunlar eksik: {missing_tr}. "
            f"Mevcut sütunlar: {', '.join(columns)}"
        )

    for alias in STOK_ALIASES:
        if alias in normalized:
            mapping["stok_no"] = normalized[alias]
            break

    for alias in ("tanim", "tanım", "urun_adi", "ürün_adı", "aciklama", "açıklama"):
        if alias in normalized:
            mapping["tanim"] = normalized[alias]
            break

    return mapping


def _parse_quantity(val: Any) -> float:
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan"):
        return 0.0
    s = s.split()[0]
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s:
        return 0.0
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return 0.0


def _normalize_stok(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val).rstrip("0").rstrip(".")
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan"):
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


class ExcelStockRepository(StockRepository):
    """
    ExcelInventoryCache — Excel tek seferlik okunur, RAM indeksleri paylaşılır.

    İndeksler:
      _by_etiket          → barkod/etiket O(1)
      _by_shelf           → raf → satır listesi O(1)
      _shelf_items_cache  → raf → API satır dict listesi (önceden hesaplanmış)
      _shelf_order        → sıralı raf listesi
      _by_stok            → stok kodu → etiket listesi O(1)
    """

    def __init__(self) -> None:
        self._lock = threading.RLock()
        self._rows: list[StockRow] = []
        self._by_etiket: dict[str, EtiketInfo] = {}
        self._by_shelf: dict[str, list[StockRow]] = {}
        self._shelf_items_cache: dict[str, list[dict[str, Any]]] = {}
        self._shelf_order: list[str] = []
        self._by_stok: dict[str, list[str]] = {}
        self._metadata: dict[str, Any] = {}
        self._columns: list[str] = []
        self._etiket_format_pattern = default_etiket_pattern()

    def load_from_excel(self, file_path: str) -> dict[str, Any]:
        with self._lock:
            return self._load_from_excel_locked(file_path)

    def _load_from_excel_locked(self, file_path: str) -> dict[str, Any]:
        path = Path(file_path)
        if not path.is_file():
            raise ValueError(f"Excel dosyası bulunamadı: {path}")

        file_bytes = path.read_bytes()
        file_hash = hashlib.sha256(file_bytes).hexdigest()

        if (
            file_hash == self._metadata.get("file_hash")
            and self._by_etiket
            and str(path) == self._metadata.get("file_path")
        ):
            logger.info(
                "Excel RAM cache güncel — disk okunmadı: %s (%d etiket)",
                path.name,
                len(self._by_etiket),
            )
            return dict(self._metadata)

        t0 = time.perf_counter()
        logger.info("Excel diskten RAM cache'e yükleniyor: %s", path)

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("Excel dosyasında sayfa bulunamadı.")

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            wb.close()
            raise ValueError("Excel dosyası boş.")

        columns = [str(c).strip() if c is not None else "" for c in header_row]
        col_map = _resolve_columns(columns)
        etiket_idx = columns.index(col_map["etiket"])
        shelf_idx = columns.index(col_map["depo"])
        qty_idx = columns.index(col_map["miktar"])
        stok_idx = columns.index(col_map["stok_no"]) if "stok_no" in col_map else None
        mapped_cols = set(col_map.values())
        extra_indices = [
            (i, col) for i, col in enumerate(columns) if col not in mapped_cols
        ]

        rows: list[StockRow] = []
        by_etiket: dict[str, EtiketInfo] = {}

        for excel_row_index, row in enumerate(rows_iter, start=2):
            if not row or all(c is None for c in row):
                continue

            etiket_raw = row[etiket_idx] if etiket_idx < len(row) else None
            etiket = normalize_etiket(etiket_raw)
            if not etiket:
                continue

            shelf_raw = row[shelf_idx] if shelf_idx < len(row) else None
            shelf = normalize_depo(shelf_raw)

            qty_raw = row[qty_idx] if qty_idx < len(row) else 0
            quantity = _parse_quantity(qty_raw)

            stok_no = ""
            if stok_idx is not None and stok_idx < len(row):
                stok_no = _normalize_stok(row[stok_idx])

            extra: dict[str, Any] = {}
            for idx, col in extra_indices:
                if idx < len(row):
                    val = row[idx]
                    if val is not None and str(val).strip().lower() not in ("none", "nan"):
                        extra[col] = val
            if stok_no:
                extra["Stok No"] = stok_no
            if shelf:
                extra["Depo"] = shelf
            tanim_col = col_map.get("tanim")
            if tanim_col and tanim_col in columns:
                ti = columns.index(tanim_col)
                if ti < len(row) and row[ti] is not None:
                    extra["Tanım"] = row[ti]

            rows.append(
                StockRow(
                    etiket=etiket,
                    shelf=shelf,
                    quantity=quantity,
                    row_index=excel_row_index,
                    extra=extra,
                )
            )

            if etiket not in by_etiket:
                by_etiket[etiket] = EtiketInfo(
                    etiket=etiket,
                    shelves={},
                    total_expected=0.0,
                    is_unassigned=True,
                    extra=extra.copy(),
                )

            info = by_etiket[etiket]
            if shelf:
                info.shelves[shelf] = info.shelves.get(shelf, 0.0) + quantity
                info.is_unassigned = False

            info.total_expected += quantity
            if extra and not info.extra:
                info.extra = extra.copy()

        wb.close()

        fmt = infer_etiket_format(by_etiket.keys())
        etiket_format_pattern = compile_etiket_pattern(fmt["pattern"])
        unassigned = sum(1 for r in by_etiket.values() if r.is_unassigned)
        unique_shelves = {normalize_depo(r.shelf) for r in rows if normalize_depo(r.shelf)}

        by_shelf, shelf_items_cache, shelf_order, by_stok = self._build_indexes(rows)

        elapsed_ms = round((time.perf_counter() - t0) * 1000, 1)

        self._rows = rows
        self._by_etiket = by_etiket
        self._by_shelf = by_shelf
        self._shelf_items_cache = shelf_items_cache
        self._shelf_order = shelf_order
        self._by_stok = by_stok
        self._columns = columns
        self._etiket_format_pattern = etiket_format_pattern
        self._metadata = {
            "filename": path.name,
            "file_path": str(path),
            "file_hash": file_hash,
            "row_count": len(rows),
            "etiket_count": len(by_etiket),
            "reference_count": len(by_etiket),
            "shelf_count": len(unique_shelves),
            "unassigned_count": unassigned,
            "columns": columns,
            "miktar_column": col_map.get("miktar"),
            "stok_column": col_map.get("stok_no"),
            "depo_column": col_map.get("depo"),
            "etiket_digit_length": fmt["digit_length"],
            "etiket_pattern": fmt["pattern"],
            "cache_index_ms": elapsed_ms,
            "cache_mode": "ram",
        }

        logger.info(
            "Excel RAM cache hazır: %d satır, %d etiket, %d raf, %d stok kodu — %s ms",
            len(rows),
            len(by_etiket),
            len(shelf_order),
            len(by_stok),
            elapsed_ms,
        )
        logger.debug("Sütun eşlemesi: %s", col_map)
        return dict(self._metadata)

    @staticmethod
    def _build_indexes(
        rows: list[StockRow],
    ) -> tuple[
        dict[str, list[StockRow]],
        dict[str, list[dict[str, Any]]],
        list[str],
        dict[str, list[str]],
    ]:
        by_shelf: dict[str, list[StockRow]] = {}
        by_stok: dict[str, list[str]] = {}
        seen_stok_etiket: set[tuple[str, str]] = set()

        for row in rows:
            shelf = normalize_depo(row.shelf)
            if shelf:
                by_shelf.setdefault(shelf, []).append(row)

            stok_no = str(row.extra.get("Stok No", "")).strip()
            if stok_no:
                key = (stok_no, row.etiket)
                if key not in seen_stok_etiket:
                    seen_stok_etiket.add(key)
                    by_stok.setdefault(stok_no, []).append(row.etiket)

        shelf_items_cache: dict[str, list[dict[str, Any]]] = {}
        for shelf, shelf_rows in by_shelf.items():
            items: list[dict[str, Any]] = []
            for row in shelf_rows:
                extra = dict(row.extra)
                extra["Depo"] = row.shelf
                items.append(
                    {
                        "line_id": f"{shelf}::R{row.row_index}",
                        "etiket": row.etiket,
                        "expected": row.quantity,
                        "extra": extra,
                    }
                )
            items.sort(key=lambda x: int(x["line_id"].split("::R")[-1]))
            shelf_items_cache[shelf] = items

        shelf_order = sorted(by_shelf.keys())
        return by_shelf, shelf_items_cache, shelf_order, by_stok

    def get_etiket(self, etiket: str) -> Optional[dict[str, Any]]:
        with self._lock:
            code = normalize_etiket(etiket)
            info = self._by_etiket.get(code)
            if not info:
                return None
            return {
                "etiket": info.etiket,
                "shelves": dict(info.shelves),
                "total_expected": info.total_expected,
                "is_unassigned": info.is_unassigned,
                "extra": dict(info.extra) if info.extra else {},
            }

    def get_etikets_by_stok(self, stok_no: str) -> list[str]:
        """Stok kodu indeksinden etiket listesi — O(1) lookup."""
        with self._lock:
            key = _normalize_stok(stok_no)
            if not key:
                return []
            return list(self._by_stok.get(key, []))

    def get_shelves(self) -> list[str]:
        with self._lock:
            return list(self._shelf_order)

    def get_shelf_items(self, shelf: str) -> list[dict[str, Any]]:
        """Raf satırları — önceden hesaplanmış cache, disk erişimi yok."""
        with self._lock:
            target = normalize_depo(shelf)
            if not target:
                return []
            cached = self._shelf_items_cache.get(target)
            if cached is not None:
                return cached
            return []

    def clear(self) -> None:
        with self._lock:
            self._rows.clear()
            self._by_etiket.clear()
            self._by_shelf.clear()
            self._shelf_items_cache.clear()
            self._shelf_order.clear()
            self._by_stok.clear()
            self._metadata.clear()
            self._columns.clear()
            self._etiket_format_pattern = default_etiket_pattern()
            logger.info("Excel RAM cache temizlendi.")

    def matches_etiket_format(self, code: str) -> bool:
        with self._lock:
            return matches_etiket_format(code, self._etiket_format_pattern)

    def get_metadata(self) -> dict[str, Any]:
        with self._lock:
            return dict(self._metadata)

    def is_loaded(self) -> bool:
        with self._lock:
            return bool(self._by_etiket)


# Geriye dönük isim — tek global örnek deps.stock_repo üzerinden kullanılır
ExcelInventoryCache = ExcelStockRepository
